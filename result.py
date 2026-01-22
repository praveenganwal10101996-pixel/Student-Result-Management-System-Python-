from openpyxl import Workbook, load_workbook
import os

FILE_NAME = "student_result.xlsx"

def Students_result():
    students_class = input("Enter Class Name (Example: Class_10): ")

    
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        
        wb.save(FILE_NAME)

    wb = load_workbook(FILE_NAME)

    # Create or load sheet
    if students_class in wb.sheetnames:
        ws = wb[students_class]
        print("Class already exists. Data will be added.")
        subjects = [cell.value for cell in ws[1][2:-2]]
    else:
        ws = wb.create_sheet(title=students_class)

        
        ws.append(["Roll No", "Name"])  

        
        subjects_input = input(
            "Enter subject names separated by comma (e.g. Maths,English,Science): "
        )
        subjects = [s.strip() for s in subjects_input.split(",")]

        for sub in subjects:
            ws.cell(row=1, column=ws.max_column + 1, value=sub)

        
        ws.cell(row=1, column=ws.max_column + 1, value="Total")
        ws.cell(row=1, column=ws.max_column + 1, value="Result")

        print(" Subjects added:", subjects)

    while True:
        roll_no = input("Enter Roll No: ")
        name = input("Enter Student Name: ")

        row_data = [roll_no, name]
        total = 0

        for sub in subjects:
            while True:
                try:
                    marks = float(input(f"Enter marks in {sub}: "))
                    break
                except ValueError:
                    print("Enter numeric marks only")
            row_data.append(marks)
            total += marks

        result = "Pass" if all(m >= 33 for m in row_data[2:]) else "Fail"

        row_data.extend([total, result])
        ws.append(row_data)

        choice = input("Add another student? (y/n): ")
        if choice.lower() == "n":
            break

    wb.save(FILE_NAME)
    print(f" Data saved in sheet: {students_class}")

def result():
    while True:
        print("\n____ Student Result System ____")
        print("1. Add Class Result")
        print("2. Exit")

        choice = input("Choose option: ")

        if choice == "1":
            Students_result()
        elif choice == "2":
            print("Program Ended")
            break
        else:
            print("Invalid choice")

result()